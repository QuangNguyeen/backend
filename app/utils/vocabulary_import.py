"""Pure parsing helpers for vocabulary CSV/XLSX import (no DB, no HTTP)."""

import csv
import io

# Maps known header aliases (lowercased) to a canonical SavedWord field.
COLUMN_ALIASES: dict[str, set[str]] = {
    "word": {"word", "term", "vocab", "vocabulary", "english", "en"},
    "meaning": {
        "meaning",
        "definition",
        "translation",
        "trans",
        "vietnamese",
        "vi",
        "nghia",
        "nghĩa",
        "ý nghĩa",
    },
    "phonetic": {"phonetic", "ipa", "pronunciation", "phiên âm"},
    "part_of_speech": {"part_of_speech", "pos", "type", "word_type", "loại từ", "từ loại"},
    "note": {"note", "notes", "comment", "ghi chú"},
    "context_sentence": {
        "context_sentence",
        "context",
        "sentence",
        "example",
        "câu ví dụ",
        "ví dụ",
    },
}


def build_header_map(headers: list[str]) -> dict[int, str]:
    """Map column indices to canonical SavedWord fields using header aliases."""
    mapping: dict[int, str] = {}
    for idx, raw in enumerate(headers):
        key = (raw or "").strip().lower()
        for field, aliases in COLUMN_ALIASES.items():
            if key == field or key in aliases:
                mapping[idx] = field
                break
    return mapping


def _parse_csv(content: bytes) -> list[dict]:
    text = content.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return []
    header_map = build_header_map(rows[0])
    out: list[dict] = []
    for cells in rows[1:]:
        record = {}
        for idx, field in header_map.items():
            if idx < len(cells):
                record[field] = (cells[idx] or "").strip()
        out.append(record)
    return out


def _parse_xlsx(content: bytes) -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return []
    header_map = build_header_map([str(h) if h is not None else "" for h in header])
    out: list[dict] = []
    for cells in rows_iter:
        record = {}
        for idx, field in header_map.items():
            if idx < len(cells) and cells[idx] is not None:
                record[field] = str(cells[idx]).strip()
        out.append(record)
    wb.close()
    return out


def parse_import_file(filename: str, content: bytes) -> list[dict]:
    name = (filename or "").lower()
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        return _parse_xlsx(content)
    return _parse_csv(content)