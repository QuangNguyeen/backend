import asyncio
import csv
import hashlib
import io
import logging
from datetime import UTC, datetime

from fastapi import APIRouter, BackgroundTasks, Depends, File, Form, Request, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy import func, select
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.deps import get_current_user
from app.core.exceptions import BadRequestError, NotFoundError
from app.database import async_session, get_db
from app.models.import_job import ImportJob
from app.models.user import User
from app.models.vocabulary import SavedWord
from app.models.word_cache import WordCache
from app.schemas.vocabulary import (
    BatchPreviewRequest,
    DueCardsResponse,
    FlashCardResponse,
    ImportJobStatus,
    ImportResultResponse,
    ReviewRequest,
    ReviewResponse,
    SavedWordResponse,
    SaveWordRequest,
    UpdateWordRequest,
    WordPreviewResponse,
)
from app.services.level_service import _get_nlp
from app.services.llm_service import (
    _TTS_FALLBACK_PREFIX,
    DictionaryResult,
    fetch_dictionary_data,
    google_translate,
)
from app.services.srs_service import calculate_next_review

MAX_IMPORT_WORDS = 100

# Maps known header aliases (lowercased) to a canonical SavedWord field.
_COLUMN_ALIASES: dict[str, set[str]] = {
    "word": {"word", "term", "vocab", "vocabulary", "english", "en"},
    "meaning": {
        "meaning",
        "definition",
        "translation",
        "trans",
        "vietnamese",
        "vi",
        "nghia",
        "nghĩa",
        "ý nghĩa",
    },
    "phonetic": {"phonetic", "ipa", "pronunciation", "phiên âm"},
    "part_of_speech": {"part_of_speech", "pos", "type", "word_type", "loại từ", "từ loại"},
    "note": {"note", "notes", "comment", "ghi chú"},
    "context_sentence": {
        "context_sentence",
        "context",
        "sentence",
        "example",
        "câu ví dụ",
        "ví dụ",
    },
}

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/vocabulary", tags=["Vocabulary"])


def _hash_context(context_sentence: str) -> str:
    return hashlib.sha256(context_sentence.encode("utf-8")).hexdigest()


def _resolve_audio_url(audio_url: str | None, request: Request) -> str | None:
    """Convert an internal TTS path to an absolute URL using the request origin."""
    if audio_url and audio_url.startswith(_TTS_FALLBACK_PREFIX):
        return str(request.base_url).rstrip("/") + audio_url
    return audio_url


_lemma_cache: dict[tuple[str, str | None], str] = {}
_LEMMA_CACHE_MAX = 5000


def _lemmatize(word: str, context_sentence: str | None = None) -> str:
    cache_key = (word, context_sentence)
    cached = _lemma_cache.get(cache_key)
    if cached is not None:
        return cached

    if "-" in word:
        _lemma_cache[cache_key] = word
        return word

    try:
        nlp = _get_nlp()

        if context_sentence:
            doc = nlp(context_sentence)
            target = word.lower()
            for token in doc:
                if token.text.lower() == target:
                    lemma = token.lemma_
                    if lemma == "-PRON-" or not lemma:
                        _lemma_cache[cache_key] = word
                        return word
                    _lemma_cache[cache_key] = lemma
                    if len(_lemma_cache) > _LEMMA_CACHE_MAX:
                        _first = next(iter(_lemma_cache))
                        _lemma_cache.pop(_first, None)
                    return lemma

        doc = nlp(word)
        if len(doc) == 0:
            _lemma_cache[cache_key] = word
            return word
        lemma = doc[0].lemma_
        if lemma == "-PRON-" or not lemma:
            _lemma_cache[cache_key] = word
            return word
        _lemma_cache[cache_key] = lemma
        if len(_lemma_cache) > _LEMMA_CACHE_MAX:
            _first = next(iter(_lemma_cache))
            _lemma_cache.pop(_first, None)
        return lemma
    except Exception as e:
        logger.warning("Lemmatization failed for word=%r: %s", word, e)
        return word


_translate_cache: dict[str, tuple[float, str | None]] = {}
_TRANSLATE_CACHE_TTL = 3600
_TRANSLATE_CACHE_MAX = 2000


async def _translate_sentence(sentence: str) -> str | None:
    """Translate with in-memory cache (same sentences reappear across users)."""
    import time

    cached = _translate_cache.get(sentence)
    if cached and (time.monotonic() - cached[0]) < _TRANSLATE_CACHE_TTL:
        return cached[1]
    result = await google_translate(sentence, target_language="vi")
    _translate_cache[sentence] = (time.monotonic(), result)
    if len(_translate_cache) > _TRANSLATE_CACHE_MAX:
        _first = next(iter(_translate_cache))
        _translate_cache.pop(_first, None)
    return result


async def _enrich_and_persist(
    saved_word_id: str,
    word_text: str,
    original_word: str,
    context_sentence: str,
    context_hash: str,
    overwrite_meaning: bool,
) -> None:
    """Background task: fetch dictionary data, update SavedWord + WordCache.

    Checks global WordCache first; on miss calls dictionaryapi.dev + Gemini.
    Context sentence translation is always fetched via Google Translate.
    """
    # Check global cache first
    phonetic: str | None = None
    vietnamese_meaning: str | None = None
    part_of_speech: str | None = None
    audio_url: str | None = None

    try:
        async with async_session() as session:
            result = await session.execute(
                select(WordCache)
                .where(
                    WordCache.word == word_text,
                    WordCache.vietnamese_meaning.isnot(None),
                )
                .limit(1)
            )
            cached = result.scalar_one_or_none()
            if cached:
                phonetic = cached.phonetic
                vietnamese_meaning = cached.vietnamese_meaning
                part_of_speech = cached.part_of_speech
                audio_url = cached.audio_url
    except Exception:
        pass

    display_meaning: str | None = vietnamese_meaning
    if not vietnamese_meaning:
        try:
            dict_data = await fetch_dictionary_data(word_text, translate_meaning=True)
        except Exception as e:
            logger.warning("Dictionary lookup crashed for word_id=%s: %s", saved_word_id, e)
            dict_data = DictionaryResult()

        phonetic = dict_data.phonetic
        vietnamese_meaning = dict_data.meaning_vi
        display_meaning = vietnamese_meaning or dict_data.meaning_en
        part_of_speech = dict_data.part_of_speech
        audio_url = dict_data.audio_url

    if not audio_url:
        from urllib.parse import quote

        audio_url = f"{_TTS_FALLBACK_PREFIX}{quote(word_text)}"

    context_translation: str | None = None
    if context_sentence:
        context_translation = await _translate_sentence(context_sentence)

    have_any = any(
        v is not None for v in (phonetic, audio_url, display_meaning, context_translation)
    )
    if not have_any:
        return

    try:
        async with async_session() as session:
            result = await session.execute(select(SavedWord).where(SavedWord.id == saved_word_id))
            saved = result.scalar_one_or_none()
            if saved is not None:
                if phonetic is not None:
                    saved.phonetic = phonetic
                if audio_url is not None:
                    saved.audio_url = audio_url
                if context_translation is not None:
                    saved.context_translation = context_translation
                if part_of_speech is not None:
                    saved.part_of_speech = part_of_speech
                if overwrite_meaning and display_meaning is not None:
                    saved.meaning = display_meaning
                await session.commit()
    except Exception as e:
        logger.exception("SavedWord update failed for word_id=%s: %s", saved_word_id, e)

    # Persist to global cache — only store actual Vietnamese, never English fallback
    try:
        async with async_session() as session:
            stmt = (
                pg_insert(WordCache)
                .values(
                    word=word_text,
                    context_hash="__global__",
                    phonetic=phonetic,
                    audio_url=audio_url,
                    vietnamese_meaning=vietnamese_meaning,
                    part_of_speech=part_of_speech,
                    context_translation=context_translation,
                )
                .on_conflict_do_nothing(
                    index_elements=["word", "context_hash"],
                )
            )
            await session.execute(stmt)
            await session.commit()
    except Exception as e:
        logger.exception("WordCache upsert failed for word=%r: %s", word_text, e)


@router.post("/save", response_model=SavedWordResponse, status_code=status.HTTP_201_CREATED)
async def save_word(
    body: SaveWordRequest,
    background_tasks: BackgroundTasks,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Save a word from dictation/quiz for SRS review (upsert).

    If the user already has this word (same lemma, not soft-deleted), the
    existing row is updated instead of creating a duplicate.

    Enrichment flow:
      - If a WordCache row exists for (word, context_hash), hydrate immediately.
      - Frontend preview fields (audio_url, phonetic, etc.) are used as fallback.
      - Otherwise, schedule a background enrichment task.
    """
    lemma = _lemmatize(body.word, body.context_sentence).lower()

    context_hash: str | None = None
    if body.context_sentence:
        context_hash = _hash_context(body.context_sentence)

    # Global word-level cache lookup
    cache_result = await db.execute(
        select(WordCache)
        .where(
            WordCache.word == lemma,
            WordCache.vietnamese_meaning.isnot(None),
        )
        .limit(1)
    )
    cached = cache_result.scalar_one_or_none()

    meaning = body.meaning
    phonetic = body.phonetic
    audio_url = body.audio_url
    context_translation = body.context_translation
    part_of_speech = body.part_of_speech

    if cached is not None:
        phonetic = phonetic or cached.phonetic
        audio_url = audio_url or cached.audio_url
        context_translation = context_translation or cached.context_translation
        part_of_speech = part_of_speech or cached.part_of_speech
        if meaning is None and cached.vietnamese_meaning is not None:
            meaning = cached.vietnamese_meaning

    # Upsert: check if user already has this word (not soft-deleted)
    existing_result = await db.execute(
        select(SavedWord).where(
            SavedWord.user_id == current_user.id,
            SavedWord.word == lemma,
            SavedWord.deleted_at.is_(None),
        )
    )
    existing = existing_result.scalar_one_or_none()

    if existing is not None:
        if body.context_sentence is not None:
            existing.context_sentence = body.context_sentence
        if body.video_id is not None:
            existing.video_id = body.video_id
        if body.audio_start_time is not None:
            existing.audio_start_time = body.audio_start_time
        if meaning is not None:
            existing.meaning = meaning
        if body.note is not None:
            existing.note = body.note
        if phonetic is not None:
            existing.phonetic = phonetic
        if audio_url is not None:
            existing.audio_url = audio_url
        if context_translation is not None:
            existing.context_translation = context_translation
        if part_of_speech is not None:
            existing.part_of_speech = part_of_speech
        existing.source = body.source
        await db.commit()
        await db.refresh(existing)
        word = existing
    else:
        word = SavedWord(
            user_id=current_user.id,
            word=lemma,
            video_id=body.video_id,
            context_sentence=body.context_sentence,
            audio_start_time=body.audio_start_time,
            meaning=meaning,
            note=body.note,
            source=body.source,
            phonetic=phonetic,
            audio_url=audio_url,
            context_translation=context_translation,
            part_of_speech=part_of_speech,
            next_review_at=datetime.now(UTC),
        )
        db.add(word)
        await db.commit()
        await db.refresh(word)

    if cached is None and body.context_sentence and context_hash is not None:
        background_tasks.add_task(
            _enrich_and_persist,
            saved_word_id=word.id,
            word_text=lemma,
            original_word=body.word,
            context_sentence=body.context_sentence,
            context_hash=context_hash,
            overwrite_meaning=body.meaning is None,
        )

    return word


@router.get("", response_model=list[SavedWordResponse])
async def list_words(
    video_id: str | None = None,
    limit: int = 50,
    offset: int = 0,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """List all saved words for the current user."""
    query = select(SavedWord).where(
        SavedWord.user_id == current_user.id,
        SavedWord.deleted_at.is_(None),
    )
    if video_id:
        query = query.where(SavedWord.video_id == video_id)
    query = query.order_by(SavedWord.created_at.desc()).limit(limit).offset(offset)

    result = await db.execute(query)
    return result.scalars().all()


@router.get("/due", response_model=DueCardsResponse)
async def get_due_cards(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Get FlashCards due for review today."""
    now = datetime.now(UTC)
    result = await db.execute(
        select(SavedWord)
        .where(
            SavedWord.user_id == current_user.id,
            SavedWord.deleted_at.is_(None),
            SavedWord.next_review_at <= now,
        )
        .order_by(SavedWord.next_review_at)
        .limit(50)
    )
    cards = result.scalars().all()

    # Count total due
    count_result = await db.execute(
        select(func.count()).where(
            SavedWord.user_id == current_user.id,
            SavedWord.deleted_at.is_(None),
            SavedWord.next_review_at <= now,
        )
    )
    total_due = count_result.scalar() or 0

    return DueCardsResponse(
        cards=[FlashCardResponse.model_validate(c) for c in cards],
        total_due=total_due,
    )


@router.get("/preview", response_model=WordPreviewResponse)
async def preview_word(
    request: Request,
    word: str,
    context: str | None = None,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Return enrichment data for a word without saving it.

    Fast path: global WordCache lookup by word (ignores context_hash).
    Slow path: parallel dictionaryapi.dev + Gemini, then cache result.
    Context translation is always handled separately via Google Translate.
    """
    lemma = _lemmatize(word, context).lower()

    # ── Parallel DB queries: is_saved + cache lookup ──
    saved_result, cache_result = await asyncio.gather(
        db.execute(
            select(SavedWord.id)
            .where(
                SavedWord.user_id == current_user.id,
                SavedWord.word == lemma,
                SavedWord.deleted_at.is_(None),
            )
            .limit(1)
        ),
        db.execute(
            select(WordCache)
            .where(
                WordCache.word == lemma,
                WordCache.vietnamese_meaning.isnot(None),
            )
            .limit(1)
        ),
    )
    is_saved = saved_result.scalar() is not None
    cached = cache_result.scalar_one_or_none()

    # ── Fast path: cache hit → return immediately ──
    if cached:
        context_translation = cached.context_translation
        if not context_translation and context:
            context_translation = await _translate_sentence(context)
        return WordPreviewResponse(
            word=lemma,
            phonetic=cached.phonetic,
            meaning=cached.vietnamese_meaning,
            audio_url=_resolve_audio_url(cached.audio_url, request),
            context_translation=context_translation,
            is_saved=is_saved,
            part_of_speech=cached.part_of_speech,
        )

    # ── Slow path: parallel Gemini + dictionaryapi.dev + context translation ──
    tasks: list[asyncio.Task] = []
    tasks.append(asyncio.ensure_future(fetch_dictionary_data(lemma, translate_meaning=True)))
    if context:
        tasks.append(asyncio.ensure_future(_translate_sentence(context)))

    results = await asyncio.gather(*tasks, return_exceptions=True)

    dict_data = results[0] if not isinstance(results[0], Exception) else DictionaryResult()
    if isinstance(dict_data, Exception):
        logger.warning("Dictionary lookup failed for word=%r: %s", word, dict_data)
        dict_data = DictionaryResult()

    context_translation: str | None = None
    if context and len(results) > 1 and not isinstance(results[1], Exception):
        context_translation = results[1]

    phonetic = dict_data.phonetic
    vietnamese_meaning = dict_data.meaning_vi
    display_meaning = vietnamese_meaning or dict_data.meaning_en
    part_of_speech = dict_data.part_of_speech
    audio_url = dict_data.audio_url
    if not audio_url:
        from urllib.parse import quote

        audio_url = f"{_TTS_FALLBACK_PREFIX}{quote(lemma)}"

    # Persist to global cache — only store actual Vietnamese as vietnamese_meaning
    try:
        cache_values = dict(
            word=lemma,
            context_hash="__global__",
            phonetic=phonetic,
            audio_url=audio_url,
            vietnamese_meaning=vietnamese_meaning,
            part_of_speech=part_of_speech,
        )
        if context_translation is not None:
            cache_values["context_translation"] = context_translation
        stmt = (
            pg_insert(WordCache)
            .values(
                **cache_values,
            )
            .on_conflict_do_nothing(index_elements=["word", "context_hash"])
        )
        await db.execute(stmt)
        await db.commit()
    except Exception as e:
        logger.warning("WordCache upsert failed in preview for word=%r: %s", lemma, e)

    return WordPreviewResponse(
        word=lemma,
        phonetic=phonetic,
        meaning=display_meaning,
        audio_url=_resolve_audio_url(audio_url, request),
        context_translation=context_translation,
        is_saved=is_saved,
        part_of_speech=part_of_speech,
    )


@router.post("/preview/batch", response_model=dict[str, WordPreviewResponse])
async def preview_words_batch(
    request: Request,
    body: BatchPreviewRequest,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Batch preview: warm caches for multiple words in one round-trip.

    Returns a dict mapping each input word to its preview data.
    Runs all dictionary lookups in parallel for maximum throughput.
    """
    if not body.words:
        return {}
    unique_words = list(dict.fromkeys(w.strip().lower() for w in body.words[:50]))

    lemmas = {w: _lemmatize(w, body.context).lower() for w in unique_words}
    unique_lemmas = list(set(lemmas.values()))

    saved_result = await db.execute(
        select(SavedWord.word).where(
            SavedWord.user_id == current_user.id,
            SavedWord.word.in_(unique_lemmas),
            SavedWord.deleted_at.is_(None),
        )
    )
    saved_set = set(saved_result.scalars().all())

    cache_result = await db.execute(
        select(WordCache).where(
            WordCache.word.in_(unique_lemmas),
            WordCache.vietnamese_meaning.isnot(None),
        )
    )
    cache_map = {row.word: row for row in cache_result.scalars().all()}

    uncached_lemmas = [lemma for lemma in unique_lemmas if lemma not in cache_map]
    dict_results: dict[str, DictionaryResult] = {}
    if uncached_lemmas:
        tasks = [
            fetch_dictionary_data(lemma, translate_meaning=True)
            for lemma in uncached_lemmas
        ]
        results = await asyncio.gather(*tasks, return_exceptions=True)
        for lemma, res in zip(uncached_lemmas, results):
            if isinstance(res, Exception):
                dict_results[lemma] = DictionaryResult()
            else:
                dict_results[lemma] = res

    ctx_translation = None
    if body.context:
        ctx_translation = await _translate_sentence(body.context)

    out: dict[str, WordPreviewResponse] = {}
    for w in unique_words:
        lemma = lemmas[w]
        cached = cache_map.get(lemma)
        if cached:
            out[w] = WordPreviewResponse(
                word=lemma,
                phonetic=cached.phonetic,
                meaning=cached.vietnamese_meaning,
                audio_url=_resolve_audio_url(cached.audio_url, request),
                context_translation=cached.context_translation or ctx_translation,
                is_saved=lemma in saved_set,
                part_of_speech=cached.part_of_speech,
            )
        else:
            dr = dict_results.get(lemma, DictionaryResult())
            audio_url = dr.audio_url
            if not audio_url:
                from urllib.parse import quote

                audio_url = f"{_TTS_FALLBACK_PREFIX}{quote(lemma)}"
            out[w] = WordPreviewResponse(
                word=lemma,
                phonetic=dr.phonetic,
                meaning=dr.meaning_vi or dr.meaning_en,
                audio_url=_resolve_audio_url(audio_url, request),
                context_translation=ctx_translation,
                is_saved=lemma in saved_set,
                part_of_speech=dr.part_of_speech,
            )
            # Persist only Vietnamese to WordCache (avoid poisoning with English)
            try:
                cache_vals = dict(
                    word=lemma,
                    context_hash="__global__",
                    phonetic=dr.phonetic,
                    audio_url=audio_url,
                    vietnamese_meaning=dr.meaning_vi,
                    part_of_speech=dr.part_of_speech,
                )
                if ctx_translation is not None:
                    cache_vals["context_translation"] = ctx_translation
                stmt = (
                    pg_insert(WordCache)
                    .values(**cache_vals)
                    .on_conflict_do_nothing(
                        index_elements=["word", "context_hash"],
                    )
                )
                await db.execute(stmt)
            except Exception:
                pass
    try:
        await db.commit()
    except Exception:
        pass
    return out


@router.get("/tts/{word}")
async def text_to_speech(word: str):
    """Generate pronunciation audio using edge-tts (async, in-memory, no disk I/O)."""
    import edge_tts

    buf = io.BytesIO()
    try:
        communicate = edge_tts.Communicate(word, "en-US-AriaNeural", volume="+50%")
        async for chunk in communicate.stream():
            if chunk["type"] == "audio":
                buf.write(chunk["data"])
    except Exception as e:
        logger.warning("edge-tts failed for word=%r: %s", word, e)
        return StreamingResponse(
            io.BytesIO(b""),
            status_code=500,
            media_type="text/plain",
        )

    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="audio/mpeg",
        headers={"Cache-Control": "public, max-age=86400"},
    )


@router.post("/{word_id}/review", response_model=ReviewResponse)
async def review_word(
    word_id: str,
    body: ReviewRequest,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Submit a FlashCard review with SM-2 quality rating (1-5)."""
    result = await db.execute(
        select(SavedWord).where(
            SavedWord.id == word_id,
            SavedWord.user_id == current_user.id,
        )
    )
    word = result.scalar_one_or_none()
    if not word:
        raise NotFoundError("Word not found")

    # Calculate next review using SM-2
    srs = calculate_next_review(
        quality=body.quality,
        repetitions=word.repetitions,
        ease_factor=word.ease_factor,
        interval_days=word.interval_days,
    )

    # Update word
    word.repetitions = srs.repetitions
    word.ease_factor = srs.ease_factor
    word.interval_days = srs.interval_days
    word.next_review_at = srs.next_review_at
    word.last_reviewed_at = datetime.now(UTC)

    await db.commit()

    return ReviewResponse(
        word_id=word.id,
        next_review_at=srs.next_review_at,
        interval_days=srs.interval_days,
        ease_factor=srs.ease_factor,
        repetitions=srs.repetitions,
    )


@router.patch("/{word_id}", response_model=SavedWordResponse)
async def update_word(
    word_id: str,
    body: UpdateWordRequest,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Update meaning or note for a saved word."""
    result = await db.execute(
        select(SavedWord).where(
            SavedWord.id == word_id,
            SavedWord.user_id == current_user.id,
        )
    )
    word = result.scalar_one_or_none()
    if not word:
        raise NotFoundError("Word not found")

    if body.meaning is not None:
        word.meaning = body.meaning
    if body.note is not None:
        word.note = body.note

    await db.commit()
    await db.refresh(word)
    return word


@router.delete("/{word_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_word(
    word_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Soft delete a saved word (SRS 7.2.7)."""
    result = await db.execute(
        select(SavedWord).where(
            SavedWord.id == word_id,
            SavedWord.user_id == current_user.id,
            SavedWord.deleted_at.is_(None),
        )
    )
    word = result.scalar_one_or_none()
    if not word:
        raise NotFoundError("Word not found")

    word.deleted_at = datetime.now(UTC)
    await db.commit()


@router.get("/export")
async def export_words(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Export all saved words as CSV."""
    result = await db.execute(
        select(SavedWord)
        .where(SavedWord.user_id == current_user.id, SavedWord.deleted_at.is_(None))
        .order_by(SavedWord.created_at.desc())
    )
    words = result.scalars().all()

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["word", "meaning", "phonetic", "part_of_speech", "note", "context_sentence"])
    for w in words:
        writer.writerow(
            [
                w.word,
                w.meaning or "",
                w.phonetic or "",
                w.part_of_speech or "",
                w.note or "",
                w.context_sentence or "",
            ]
        )

    buf.seek(0)
    return StreamingResponse(
        io.BytesIO(buf.getvalue().encode("utf-8")),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=vocabulary.csv"},
    )


def _build_header_map(headers: list[str]) -> dict[int, str]:
    """Map column indices to canonical SavedWord fields using header aliases."""
    mapping: dict[int, str] = {}
    for idx, raw in enumerate(headers):
        key = (raw or "").strip().lower()
        for field, aliases in _COLUMN_ALIASES.items():
            if key == field or key in aliases:
                mapping[idx] = field
                break
    return mapping


def _parse_csv(content: bytes) -> list[dict]:
    text = content.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return []
    header_map = _build_header_map(rows[0])
    out: list[dict] = []
    for cells in rows[1:]:
        record = {}
        for idx, field in header_map.items():
            if idx < len(cells):
                record[field] = (cells[idx] or "").strip()
        out.append(record)
    return out


def _parse_xlsx(content: bytes) -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return []
    header_map = _build_header_map([str(h) if h is not None else "" for h in header])
    out: list[dict] = []
    for cells in rows_iter:
        record = {}
        for idx, field in header_map.items():
            if idx < len(cells) and cells[idx] is not None:
                record[field] = str(cells[idx]).strip()
        out.append(record)
    wb.close()
    return out


def _parse_import_file(filename: str, content: bytes) -> list[dict]:
    name = (filename or "").lower()
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        return _parse_xlsx(content)
    return _parse_csv(content)


@router.post("/import", response_model=ImportResultResponse)
async def import_words(
    file: UploadFile = File(...),
    enrich: bool = Form(False),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Import words from a CSV or XLSX file.

    Columns are auto-detected from the header row (word/term, meaning/definition,
    phonetic/ipa, part_of_speech/pos, note, context_sentence/example). Only
    ``word`` is required. Up to 100 words may be imported at once.

    When ``enrich=True``, an ``ImportJob`` is created and a Celery task is queued
    to fill missing meaning, IPA, audio, and example sentences in the background;
    the returned ``job_id`` can be polled via ``/import/{job_id}/status``.
    """
    content = await file.read()
    try:
        rows = _parse_import_file(file.filename or "", content)
    except Exception as e:
        logger.warning("Import file parse failed: %s", e)
        raise BadRequestError("Could not parse file. Please upload a valid CSV or XLSX.")

    # Validate size against the per-import cap (count rows that have a word).
    word_rows = [r for r in rows if (r.get("word") or "").strip()]
    if len(word_rows) > MAX_IMPORT_WORDS:
        raise BadRequestError(f"Maximum {MAX_IMPORT_WORDS} words per import")

    imported = 0
    updated = 0
    errors: list[dict] = []
    existing_ids: list[str] = []
    new_words: list[SavedWord] = []

    for row_num, row in enumerate(rows, start=2):
        word_text = (row.get("word") or "").strip().lower()
        if not word_text:
            errors.append({"row": row_num, "error": "missing word"})
            continue

        existing = (
            await db.execute(
                select(SavedWord).where(
                    SavedWord.user_id == current_user.id,
                    SavedWord.word == word_text,
                    SavedWord.deleted_at.is_(None),
                )
            )
        ).scalar_one_or_none()

        if existing:
            if row.get("meaning"):
                existing.meaning = row["meaning"].strip()
            if row.get("phonetic"):
                existing.phonetic = row["phonetic"].strip()
            if row.get("note"):
                existing.note = row["note"].strip()
            if row.get("part_of_speech"):
                existing.part_of_speech = row["part_of_speech"].strip()
            if row.get("context_sentence"):
                existing.context_sentence = row["context_sentence"].strip()
            updated += 1
            if enrich:
                existing_ids.append(existing.id)
        else:
            new_word = SavedWord(
                user_id=current_user.id,
                word=word_text,
                meaning=(row.get("meaning") or "").strip() or None,
                phonetic=(row.get("phonetic") or "").strip() or None,
                note=(row.get("note") or "").strip() or None,
                part_of_speech=(row.get("part_of_speech") or "").strip() or None,
                context_sentence=(row.get("context_sentence") or "").strip() or None,
                source="csv_import",
                next_review_at=datetime.now(UTC),
            )
            db.add(new_word)
            imported += 1
            if enrich:
                new_words.append(new_word)

    await db.commit()

    # IDs of newly-inserted rows are only populated after the flush/commit above.
    new_word_ids = existing_ids + [w.id for w in new_words]

    job_id: str | None = None
    enrich_queued = False
    if enrich and new_word_ids:
        job = ImportJob(
            user_id=current_user.id,
            status="pending",
            total_words=len(new_word_ids),
            phase="meanings",
        )
        db.add(job)
        await db.commit()
        await db.refresh(job)
        job_id = job.id
        try:
            from app.tasks.vocabulary_enrichment import enrich_imported_words

            enrich_imported_words.delay(job.id, new_word_ids, True)
            enrich_queued = True
        except Exception as e:
            logger.error("Failed to queue enrichment task for job %s: %s", job.id, e)
            job.status = "failed"
            job.error = "Could not queue enrichment task"
            await db.commit()

    return ImportResultResponse(
        job_id=job_id,
        imported=imported,
        updated=updated,
        errors=errors,
        total_words=len(word_rows),
        enrich_queued=enrich_queued,
    )


@router.get("/import/template")
async def download_template():
    """Download a sample CSV template with headers and example rows."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["word", "meaning", "phonetic", "part_of_speech", "note", "context_sentence"])
    writer.writerow(["apple", "quả táo", "/ˈæp.əl/", "noun", "", "I eat an apple every day."])
    writer.writerow(["resilient", "kiên cường", "/rɪˈzɪl.i.ənt/", "adjective", "", ""])
    writer.writerow(
        ["contribute", "đóng góp", "/kənˈtrɪb.juːt/", "verb", "", "He contributes to the team."]
    )
    buf.seek(0)
    return StreamingResponse(
        io.BytesIO(buf.getvalue().encode("utf-8-sig")),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=vocabulary_template.csv"},
    )


@router.get("/import/{job_id}/status", response_model=ImportJobStatus)
async def get_import_status(
    job_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Poll enrichment progress for an import job."""
    result = await db.execute(
        select(ImportJob).where(
            ImportJob.id == job_id,
            ImportJob.user_id == current_user.id,
        )
    )
    job = result.scalar_one_or_none()
    if not job:
        raise NotFoundError("Import job not found")

    total = job.total_words or 0
    enriched = min(job.enriched_count or 0, total) if total else 0
    progress_pct = (
        int(round((enriched / total) * 100)) if total else (100 if job.status == "completed" else 0)
    )
    if job.status == "completed":
        progress_pct = 100

    return ImportJobStatus(
        job_id=job.id,
        status=job.status,
        total=total,
        enriched=enriched,
        phase=job.phase,
        progress_pct=progress_pct,
        error=job.error,
    )
